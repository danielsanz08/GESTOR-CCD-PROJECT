from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.conf import settings
from django.utils import timezone
from django.utils.timezone import localtime, make_aware, now
from django.utils.dateparse import parse_date
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from django.utils.html import escape
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string, get_template
from papeleria.models import Articulo, PedidoArticulo, Pedido, Devolucion
from papeleria.forms import LoginForm, ArticuloForm, ArticuloEditForm, PedidoArticuloForm, DevolucionForm
from libreria.models import CustomUser
import uuid
from django.contrib.staticfiles import finders
import os
from io import BytesIO
from datetime import datetime, timedelta
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from xhtml2pdf import pisa
User = get_user_model()
def wrap_text(text, max_len=19):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  
    return '\n'.join(parts)


def wrap_text_p(text, max_len=26
                ):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-' 
    return '\n'.join(parts)

def draw_table_on_canvas(canvas, doc):
    watermark_path = finders.find('imagen/LOGO.png')
    if watermark_path:
        canvas.saveState()
        canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.drawImage(watermark_path,
                         x=(doc.pagesize[0] - 600) / 2,
                         y=(doc.pagesize[1] - 600) / 2,
                         width=600, height=600, mask='auto')
        canvas.restoreState()

#VALIDAR CORREO
def validar_datos(request):
    email = request.GET.get('email', None)
    
    errores = {}

   
    if email and CustomUser.objects. filter(email=email).exists():
        errores['email'] = 'El email ya está en uso.'

    
    return JsonResponse(errores if errores else {'valid': True})

#LOGIN Y LOGOUT
def login_papeleria(request):
    if request.method == 'POST':
        print("POST recibido")
        print("Datos enviados:", request.POST)

        email = request.POST.get('email')
        password = request.POST.get('password')

        print("Email:", email)
        print("Password:", password)

        if not email or not password:
            messages.error(request, "Todos los campos son obligatorios.")
            return render(request, 'login_pap/login_pap.html')

        user = authenticate(request, email=email, password=password)

        if user is not None:
            print("Usuario autenticado:", user.email)
            if user.is_active and getattr(user, 'acceso_pap', False):
                # Cierra sesión previa si existe
                if hasattr(user, 'session_key') and user.session_key:
                    Session.objects.filter(session_key=user.session_key).delete()

                login(request, user)  # Crea nueva sesión automáticamente

                # Asigna nueva información de sesión
                session_key = request.session.session_key
                token = str(uuid.uuid4())
                request.session['session_token'] = token

                user.session_key = session_key
                user.session_token = token
                user.save()

                messages.success(request, "¡Inicio de sesión exitoso!")
                return redirect('papeleria:index_pap')
            else:
                messages.error(request, "No tienes permiso para acceder a este módulo.")
                print("Usuario sin acceso al módulo.")
        else:
            messages.error(request, "Credenciales inválidas.")
            print("Autenticación fallida: usuario no encontrado o contraseña incorrecta.")
    else:
        print("Método no POST:", request.method)

    return render(request, 'login_pap/login_pap.html')

@never_cache
@login_required(login_url='/acceso_denegado/')
def logout_view(request):
    request.session.flush() 
    messages.success(request, "Has cerrado sesión correctamente.")
    response = redirect('libreria:inicio')
    response.delete_cookie('sessionid')
    return response

#INDEX DE PAPELERIA
@never_cache
@login_required(login_url='/acceso_denegado/')
def index_pap(request):
    breadcrumbs = [{'name': 'Inicio', 'url': '/index_pap'}]
    mostrar_alerta = True 
    bajo_stock = True      
    es_papeleria = True   
    context = {
        'breadcrumbs': breadcrumbs,
        'es_papeleria': es_papeleria,
    }

    return render(request, 'index_pap/index_pap.html', context)

#VIEWS DE ARTICULOS
@login_required(login_url='/acceso_denegado/')
def crear_articulo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Crear artículo', 'url': reverse('papeleria:crear_articulo')},
    ]

    if request.method == 'POST':
        form = ArticuloForm(request.POST)
        if form.is_valid():
            try:
                articulo = form.save(commit=False)
                articulo.registrado_por = request.user
                articulo.save()
                messages.success(request, '¡Artículo creado exitosamente!')
                return redirect('papeleria:listar_articulo')
            except Exception as e:
                messages.error(request, f'Error al crear el artículo: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = ArticuloForm()

    return render(request, 'articulo/crear_articulo.html', {
        'form': form,
        'breadcrumbs': breadcrumbs
    })

@login_required(login_url='/acceso_denegado/')
def editar_articulo(request, articulo_id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Lista de articulos', 'url': reverse('papeleria:listar_articulo')}, 
        {'name': 'Editar artículo', 'url': reverse('papeleria:editar_articulo', kwargs={'articulo_id': articulo_id})},
    ]

    articulo = get_object_or_404(Articulo, id=articulo_id)

    if request.method == 'POST':
        form = ArticuloEditForm(request.POST, instance=articulo)
        if form.is_valid():
            try:
                articulo_editado = form.save(commit=False)
                articulo_editado.registrado_por = request.user
                articulo_editado.save()
                
                messages.success(request, f'¡El artículo "{articulo.nombre}" actualizado correctamente!')
                return redirect('papeleria:listar_articulo')
            except Exception as e:
                messages.error(request, f'Error al guardar los cambios: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = ArticuloEditForm(instance=articulo)

    return render(request, 'articulo/editar_articulo.html', {
        'form': form,
        'articulo': articulo,
        'breadcrumbs': breadcrumbs
    })

@login_required(login_url='/acceso_denegado/')
def listar_articulo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listar artículos', 'url': reverse('papeleria:listar_articulo')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    articulos = Articulo.objects.select_related('registrado_por').order_by('nombre')

    if query:
        articulos = articulos.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            articulos = articulos.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            articulos = Articulo.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            articulos = articulos.filter(fecha_registro__lt=fecha_fin)
        except ValueError:
            articulos = Articulo.objects.none()

    paginator = Paginator(articulos, 4)
    page_number = request.GET.get('page')
    articulos_page = paginator.get_page(page_number)

    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}' 

    return render(request, 'articulo/listar_articulo.html', {
        'articulos': articulos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

def buscar_articulo(request):
    query = request.GET.get('q', '').strip()
    if query:
        articulos = Articulo.objects.filter(nombre__icontains=query).values(
            'id', 'nombre', 'marca', 'observacion', 'precio', 'registrado_por', 'fecha_formateada', 'proveedor'
        )
        return JsonResponse(list(articulos), safe=False)
    return JsonResponse([], safe=False)

@login_required(login_url='/acceso_denegado/')
def eliminar_articulo(request, id):
    articulo = Articulo.objects.get(id=id)
    if request.method == 'POST':
        articulo.delete()
        messages.success(request, "Artículo eliminado correctamente.")
    return redirect('papeleria:listar_articulo')

@login_required(login_url='/acceso_denegado/')
def lista_stock_bajo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Stock bajo', 'url': reverse('papeleria:lista_bajo_stock')}, 
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    articulos_bajo_stock = Articulo.objects.filter(cantidad__lt=10)

    if query:
        articulos_bajo_stock = articulos_bajo_stock.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            articulos_bajo_stock = articulos_bajo_stock.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            articulos_bajo_stock = Articulo.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            articulos_bajo_stock = articulos_bajo_stock.filter(fecha_registro__lt=fecha_fin)
        except ValueError:
            articulos_bajo_stock = Articulo.objects.none()

    bajo_stock = articulos_bajo_stock.exists()
    nombres_bajo_stock = [art.nombre for art in articulos_bajo_stock]

    
    paginator = Paginator(articulos_bajo_stock, 4)
    page_number = request.GET.get('page')
    articulos_page = paginator.get_page(page_number)

    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'  

    return render(request, 'articulo/bajo_stock.html', {
        'articulos': articulos_page,
        'bajo_stock': bajo_stock,
        'nombres_bajo_stock': nombres_bajo_stock,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

def verificar_nombre_articulo(request):
    nombre = request.GET.get('nombre', '')
    existe = Articulo.objects.filter(nombre=nombre).exists()
    return JsonResponse({'existe': existe})

@login_required(login_url='/acceso_denegado/')
def estadisticas_articulos(request):
    articulos = Articulo.objects.all().order_by('-cantidad') 
    total_cantidad = articulos.aggregate(total=Sum('cantidad'))['total'] or 0
    return render(request, 'estadisticas/estadisticas_articulos.html', {
        'articulos': articulos,
        'total_cantidad': total_cantidad,
    })


@login_required(login_url='/acceso_denegado/')
def graficas_articulos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadisticas', 'url': reverse('papeleria:index_estadistica')}, 
        {'name': 'Grafico de articulos', 'url': reverse('papeleria:graficas_articulos')}, 
    ]
    articulos = Articulo.objects.all()
    
    nombres = [f"{art.nombre} {art.tipo}" for art in articulos]
    cantidades = [art.cantidad for art in articulos]

    return render(request, 'estadisticas/grafica_articulos.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })

@login_required(login_url='/acceso_denegado/')
def grafica_bajo_Stock(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadisticas', 'url': reverse('papeleria:index_estadistica')}, 
        {'name': 'Grafico de bajo stock', 'url': reverse('papeleria:grafica_bajoStock')}, 
    ]
    articulos = Articulo.objects.filter(cantidad__lt=10)
    nombres = [f"{art.nombre} {art.tipo}" for art in articulos]
    cantidades = [art.cantidad for art in articulos]
    return render(request, 'estadisticas/grafica_bajoStock.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })

def get_articulos_filtrados(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    articulos = Articulo.objects.select_related('registrado_por').order_by('id')


    if query:
        articulos = articulos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        articulos = articulos.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return articulos

@login_required(login_url='/acceso_denegado/')
def reporte_articulo_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de Artículos CCD"
    doc.author = "CCD"
    doc.subject = "Listado de artículos"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph("REPORTE DE ARTÍCULOS", styles["Title"])
    elements.append(titulo)

 
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
       ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha"],
        ["Cámara de comercio de Duitama", "Nit: 891855025", "gestiondocumental@ccduitama.org.co", f"{fecha_actual}"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        wrap_text(usuario.username),
        wrap_text(usuario.email),
        wrap_text(getattr(usuario, 'role', 'No definido')),
        wrap_text(getattr(usuario, 'cargo', 'No definido')),
])

    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    
    articulos_filtrados = get_articulos_filtrados(request)

    if not articulos_filtrados.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron artículos.", centered_style)
        elements.append(no_results)
    else:

        data_articulos = [["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]]

        for articulo in articulos_filtrados:
            data_articulos.append([
                wrap_text(str(articulo.id)),
                wrap_text(articulo.nombre),
                wrap_text(articulo.marca),
                wrap_text(articulo.tipo),
                wrap_text("{:,}".format(articulo.precio)),
                wrap_text(str(articulo.cantidad)),
                wrap_text(articulo.observacion),
            ])

        tabla_articulos = Table(data_articulos, colWidths=[20, 140, 120, 110, 100, 80, 150])
        style_articulos = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado de articulos de papeleria CCD.pdf"'
    return response

@login_required(login_url='/acceso_denegado/')
def reporte_articulo_excel(request):
    articulos = get_articulos_filtrados(request)

    
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de Artículos CCD"

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 33

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:G1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Artículos {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    for row in [1, 2]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f"{col}3"]
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    if not articulos.exists():
        ws.merge_cells('A4:G4')
        ws['A4'] = "No se encontraron artículos."
        ws['A4'].font = Font(size=14)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A4'].border = border
    else:
        
        for articulo in articulos:
            ws.append([
                wrap_text(str(articulo.id)),
                wrap_text(articulo.nombre),
                wrap_text(articulo.marca),
                wrap_text(articulo.tipo),
                wrap_text("{:,}".format(articulo.precio)),
                wrap_text(str(articulo.cantidad)),
                str((articulo.observacion)),
            ])

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.auto_filter.ref = "A3:G3"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Listado de articulos de papeleria CCD.xlsx"'
    wb.save(response)
    return response

def obtener_articulos_bajo_stock(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    articulos = Articulo.objects.filter(cantidad__lt=10)

    if query:
        articulos = articulos.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        articulos = articulos.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return articulos

@login_required(login_url='/acceso_denegado/')
def reporte_articulo_bajo_stock_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de BAJO STOCK CCD"
    doc.author = "CCD"
    doc.subject = "Listado de bajo stock"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    
    titulo = Paragraph("REPORTE DE BAJO STOCK", styles["Title"])
    elements.append(titulo)

    
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha"],
        ["Cámara de comercio de Duitama", "Nit: 891855025", "gestiondocumental@ccduitama.org.co", f"{fecha_actual}"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        wrap_text(usuario.username),
        wrap_text(usuario.email),
        wrap_text(getattr(usuario, 'role', 'No definido')),
        wrap_text(getattr(usuario, 'cargo', 'No definido')),
])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    articulos_filtrados = obtener_articulos_bajo_stock(request)

    if not articulos_filtrados.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER
        )
        no_results = Paragraph("No se encontraron artículos con bajo stock.", centered_style)
        elements.append(Spacer(1, 20))
        elements.append(no_results)
    else:
        data_articulos = [["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]]
        
        for articulo in articulos_filtrados:
            data_articulos.append([
                wrap_text_p(str(articulo.id)),
                wrap_text_p(articulo.nombre),
                wrap_text_p(articulo.marca),
                wrap_text_p(articulo.tipo),
                wrap_text_p("{:,}".format(articulo.precio)),
                wrap_text_p(str(articulo.cantidad)),
                wrap_text_p(articulo.observacion),
            ])

        tabla_articulos = Table(data_articulos, colWidths=[20, 140, 135, 130, 90, 55, 150])
        style_articulos = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado bajo stock de papeleria CCD.pdf"'
    return response

@login_required(login_url='/acceso_denegado/')
def reporte_articulo_bajo_stock_excel(request):
    
    articulos = obtener_articulos_bajo_stock(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de Artículos CCD"

    column_widths = [10, 40, 20, 30, 20, 15, 33]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:G1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Artículos - {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if not articulos.exists():
        ws.merge_cells('A4:G4')
        cell = ws['A4']
        cell.value = "No se encontraron artículos en bajo stock."
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    else:
        for articulo in articulos:
            precio_formateado = f"${articulo.precio:,.2f}" if articulo.precio else "-"
            ws.append([
                articulo.id,
                str(articulo.nombre),
                str(articulo.marca),
                str(articulo.tipo),
                precio_formateado,
                articulo.cantidad,
                str(articulo.observacion or ""),
            ])

    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for i, cell in enumerate(row, start=1):
            cell.border = border
            if i < 7:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.auto_filter.ref = "A3:G3"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Listado articulos con bajo stock de papeleria.xlsx"'
    wb.save(response)
    return response

#PEDIDOS
@login_required(login_url='/acceso_denegado/')
def crear_pedido(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Realizar pedidos', 'url': reverse('papeleria:crear_pedido')}, 
    ]

    if request.method == 'POST':
        estado = 'Confirmado' if request.user.role == 'Administrador' else 'Pendiente'

        fecha_personalizada_str = request.POST.get('fecha_personalizada')
        if fecha_personalizada_str:
            try:
                fecha_personalizada = datetime.strptime(fecha_personalizada_str, '%Y-%m-%dT%H:%M')
                if timezone.is_naive(fecha_personalizada):
                    fecha_personalizada = make_aware(fecha_personalizada)
            except ValueError:
                fecha_personalizada = timezone.now()
        else:
            fecha_personalizada = timezone.now()

        pedido = Pedido.objects.create(
            registrado_por=request.user,
            estado=estado,
            fecha_estado=timezone.now() if estado == 'Confirmado' else None,
            fecha_pedido=fecha_personalizada  
        )

        articulos_ids = request.POST.getlist('articulo')
        tipos = request.POST.getlist('tipo_articulo')
        cantidades = request.POST.getlist('cantidad')

        area_usuario = request.user.area  
        for articulo_id, tipo, cantidad in zip(articulos_ids, tipos, cantidades):
            if articulo_id and tipo and cantidad:
                cantidad = int(cantidad)
                articulo = Articulo.objects.get(id=articulo_id)

                if estado == 'Confirmado' and articulo.cantidad < cantidad:
                    pedido.delete()
                    messages.error(request, f"No hay suficiente stock para el artículo: {articulo.nombre} (disponible: {articulo.cantidad}, solicitado: {cantidad})")
                    return redirect('papeleria:crear_pedido')

                if estado == 'Confirmado':
                    articulo.cantidad -= cantidad
                    articulo.save()

                PedidoArticulo.objects.create(
                    pedido=pedido,
                    articulo=articulo,
                    cantidad=cantidad,
                    tipo=tipo,
                    area=area_usuario,
                )

        admin_users = CustomUser.objects.filter(role='Administrador', is_active=True)
        admin_emails = [admin.email for admin in admin_users if admin.email]

        if admin_emails:
            admin_url = request.build_absolute_uri(reverse('papeleria:mis_pedidos'))
            context = {
                'usuario': request.user,
                'pedido': pedido,
                'admin_url': admin_url,
                'company_name': 'Gestor CCD',
                'articulos': PedidoArticulo.objects.filter(pedido=pedido),
                'fecha_pedido': localtime(pedido.fecha_pedido).strftime('%d/%m/%Y %H:%M')}

            html_message = render_to_string('pedidos/email_notificacion_pedido.html', context)
            
            text_message = f"""
Hola Administrador,

Se ha registrado un nuevo pedido en el sistema de Papelería:

Usuario: {request.user.username}
Rol: {request.user.get_role_display()}  # Muestra el nombre legible del rol
Área: {request.user.get_area_display()}  # Muestra el nombre legible del área
Cargo: {request.user.cargo}
ID del Pedido: {pedido.id}
Estado: {estado}
Fecha: {pedido.fecha_pedido.strftime('%d/%m/%Y %H:%M')}

Detalle de artículos:
{'\n'.join([f"- {pa.articulo.nombre} x {pa.cantidad} ({pa.tipo})" for pa in PedidoArticulo.objects.filter(pedido=pedido)])}

Por favor revisa y gestiona este pedido en el sistema: {admin_url}

Este es un mensaje automático, por favor no respondas a este correo.
"""

            try:
                subject = f"Nuevo pedido {'confirmado' if estado == 'Confirmado' else 'pendiente'} - ID: {pedido.id}"
                msg = EmailMultiAlternatives(
                    subject,
                    text_message,
                    settings.DEFAULT_FROM_EMAIL,
                    admin_emails
                )
                msg.attach_alternative(html_message, "text/html")
                msg.send()

            except TimeoutError:
                messages.error(request, "El pedido fue creado, pero no se pudo enviar la notificación por correo debido a un problema de conexión (Timeout).")
            except Exception as e:
                print(f"Error enviando correo: {e}")

        messages.success(request, f"El pedido fue registrado correctamente con estado '{estado}'.")
        return redirect('papeleria:mis_pedidos' if request.user.role == 'Empleado' else 'papeleria:mis_pedidos')

    articulos = Articulo.objects.all()
    return render(request, 'pedidos/pedidos.html', {
        'articulos': articulos,
        'breadcrumbs': breadcrumbs,
    })

@login_required(login_url='/acceso_denegado/')
def mis_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listado de pedidos', 'url': reverse('papeleria:listado_pedidos')},
    ]
    
    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    
    pedidos = Pedido.objects.filter(registrado_por=request.user).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
           Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass
            
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/mis_pedidos.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

@csrf_exempt
@require_POST
@csrf_exempt
@require_POST
def cambiar_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    nuevo_estado = request.POST.get('estado')

    if not nuevo_estado:
        messages.error(request, 'No se especificó un nuevo estado.')
        return redirect('papeleria:pedidos_pendientes')

    if nuevo_estado == 'Confirmado':
        try:
            with transaction.atomic():
                articulos_pedido = pedido.articulos.select_related('articulo').all()
                articulos_sin_stock = []

                for item in articulos_pedido:
                    articulo = item.articulo
                    if articulo.cantidad < item.cantidad:
                        articulos_sin_stock.append(
                            f"{articulo.nombre} (Solicitados: {item.cantidad}, Disponibles: {articulo.cantidad})"
                        )

                if articulos_sin_stock:
                    pedido.estado = 'Cancelado'
                    pedido.fecha_estado = timezone.now()
                    pedido.save()
                    messages.error(request, f"Pedido cancelado. Stock insuficiente: {', '.join(articulos_sin_stock)}")
                    return redirect('papeleria:pedidos_pendientes')

                for item in articulos_pedido:
                    articulo = item.articulo
                    articulo.cantidad -= item.cantidad
                    articulo.save()

                pedido.estado = 'Confirmado'
                pedido.fecha_estado = timezone.now()
                pedido.save()

                messages.success(request, 'Pedido confirmado correctamente.')

        except Exception as e:
            messages.error(request, f'Error al confirmar el pedido: {str(e)}')
            return redirect('papeleria:pedidos_pendientes')

    else:
        pedido.estado = nuevo_estado
        pedido.fecha_estado = timezone.now()
        pedido.save()
        messages.success(request, f'Estado actualizado a {nuevo_estado}.')

    usuario = pedido.registrado_por
    if usuario and usuario.email:
        articulos_lista = "\n".join(
            f"<li class='articulo-item'>{escape(item.cantidad)} {escape(item.articulo.nombre)}</li>"
            for item in pedido.articulos.all()
        )

        html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Estado de Pedido Actualizado</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .email-container {{
            background-color: #ffffff;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            overflow: hidden;
        }}
        .header {{
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .logo {{
            max-width: 180px;
            height: auto;
            margin-bottom: 15px;
        }}
        .content {{
            padding: 25px;
        }}
        .status {{
            font-size: 12px;
            font-weight: bold;
            margin: 15px 0;
            padding: 10px;
            border-radius: 5px;
            text-align: center;
        }}
        .status.Confirmado {{
            color: #155724;
        }}
        .status.Cancelado {{
            
            color: #721c24;
        }}
        .status.Pendiente {{
            color: #856404;
        }}
        .info-box {{
            background-color: #f8f9fa;
            border-left: 4px solid #007bff;
            padding: 15px;
            margin: 20px 0;
        }}
        .info-label {{
            font-weight: bold;
            color: #495057;
            display: inline-block;
            min-width: 150px;
        }}
        .articulos-list {{
            margin: 15px 0;
            padding-left: 20px;
        }}
        .articulo-item {{
            margin-bottom: 8px;
        }}
        .footer {{
            text-align: center;
            padding: 15px;
            font-size: 12px;
            color: #6c757d;
            border-top: 1px solid #e9ecef;
            background-color: #f8f9fa;
        }}
    </style>
</head>
<body>
    <div class="email-container">
        <div class="header">
            <img src="https://ccduitama.org.co/wp-content/uploads/2021/05/LOGOCCD-TRANSPARENCIA.png" alt="Logo CCD" class="logo">
            <h2>Actualización de Estado de Pedido</h2>
        </div>
        
        <div class="content">
            <p>Hola <strong>{escape(usuario.username)}</strong>,</p>
            <p>Te informamos que el estado de tu pedido del módulo de papelería ha sido actualizado:</p>
            
            <div class="info-box">
                <p><span class="info-label">Número de Pedido:</span> #{pedido.id}</p>
                <p><span class="info-label">Fecha de Actualización:</span> {pedido.fecha_estado.strftime('%d/%m/%Y %H:%M')}</p>
                <p><span class="info-label">Nuevo Estado:</span> <span class="status {pedido.estado}">{pedido.estado.upper()}</span></p>

                <h3>Detalle del Pedido:</h3>
                <ul class="articulos-list">
                    {articulos_lista}
                </ul>
            </div>
            
            <p>Para más información, puedes acceder al sistema de gestión de pedidos.</p>
        </div>
        
        <div class="footer">
            <p>Este es un mensaje automático, por favor no respondas a este correo.</p>
            <p>© {timezone.now().year} Gestor Papelería - Todos los derechos reservados</p>
        </div>
    </div>
</body>
</html>
        """

        text_content = f"""
Actualización de Estado de Pedido

Hola {usuario.username},

Tu pedido #{pedido.id} ha sido actualizado a {pedido.estado.upper()}.

Fecha de actualización: {pedido.fecha_estado.strftime('%d/%m/%Y %H:%M')}

Detalle del Pedido:
{"".join([f"- {item.cantidad} x {item.articulo.nombre}\n" for item in pedido.articulos.all()])}

Gracias por usar nuestro sistema.
© {timezone.now().year} Gestor Papelería
        """

        try:
            send_mail(
                subject=f'Pedido #{pedido.id} - Estado actualizado a {pedido.estado.upper()}',
                message=text_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email],
                html_message=html_content
            )
        except Exception as e:
            print(f"Error enviando email: {str(e)}")

    return redirect('papeleria:pedidos_pendientes')

@login_required(login_url='/acceso_denegado/')
def listado_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listado de pedidos', 'url': reverse('papeleria:listado_pedidos')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado__in=['Confirmado', 'Cancelado']).order_by('-id')

    
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass
            
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/lista_pedidos.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

@login_required(login_url='/acceso_denegado/')
def pedidos_pendientes(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Pedidos pendientes', 'url': reverse('papeleria:pedidos_pendientes')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado='Pendiente').order_by('-fecha_pedido')
    
    
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    try:
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
            
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
    except ValueError:
        pass  

    
    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/confirmar_pedido.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

@login_required(login_url='/acceso_denegado/')
def grafica_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Gráfico de estado de pedidos', 'url': reverse('papeleria:grafica_pedidos')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.all()

    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio_dt)
        except ValueError:
            pass  
    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin_dt)
        except ValueError:
            pass

    pendientes = pedidos.filter(estado='Pendiente').count()
    confirmados = pedidos.filter(estado='Confirmado').count()
    cancelados = pedidos.filter(estado='Cancelado').count()

    nombres = ['Pendiente', 'Confirmado', 'Cancelado']
    cantidades = [pendientes, confirmados, cancelados]

    return render(request, 'estadisticas/grafica_estado_pendiente.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })

@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_administrativa(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Gráfico de pedidos Administrativa', 'url': reverse('papeleria:pedidos_administrativa')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Administrativa')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None  

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(pedido__fecha_pedido__lt=fecha_fin)  # < usando < para incluir todo el día anterior
        except ValueError:
            fecha_fin = None

    pedidos_por_area_articulo = pedidos.values(
        'articulo__nombre',
        'articulo__tipo'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [
        f"{item['articulo__nombre']} ({item['articulo__tipo']})"
        if item['articulo__nombre'] and item['articulo__tipo']
        else item['articulo__nombre'] or 'Sin artículo'
        for item in pedidos_por_area_articulo
    ]

    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_area.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })

@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_rues(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos RUES', 'url': reverse('papeleria:pedidos_rues')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Registros públicos')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(pedido__fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None

    pedidos_por_area_articulo = pedidos.values(
        'articulo__nombre',
        'articulo__tipo'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [
        f"{item['articulo__nombre']} ({item['articulo__tipo']})"
        if item['articulo__nombre'] and item['articulo__tipo']
        else item['articulo__nombre'] or 'Sin artículo'
        for item in pedidos_por_area_articulo
    ]
    
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_rues.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })
@login_required(login_url='/acceso_denegado/')
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_presidencia(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos Presidencia', 'url': reverse('papeleria:pedidos_presidencia')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Presidencia')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(pedido__fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None

    pedidos_por_area_articulo = pedidos.values(
        'articulo__nombre',
        'articulo__tipo'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [
        f"{item['articulo__nombre']} ({item['articulo__tipo']})"
        if item['articulo__nombre'] and item['articulo__tipo']
        else item['articulo__nombre'] or 'Sin artículo'
        for item in pedidos_por_area_articulo
    ]
    
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_presidencia.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })
@login_required(login_url='/acceso_denegado/')
@never_cache
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_financiera(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos Financiera', 'url': reverse('papeleria:pedidos_financiera')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Financiera')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(pedido__fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None

    pedidos_por_area_articulo = pedidos.values(
        'articulo__nombre',
        'articulo__tipo'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [
        f"{item['articulo__nombre']} ({item['articulo__tipo']})"
        if item['articulo__nombre'] and item['articulo__tipo']
        else item['articulo__nombre'] or 'Sin artículo'
        for item in pedidos_por_area_articulo
    ]
    
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_financiera.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })

@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_gestion_empresarial(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos gestión empresarial', 'url': reverse('papeleria:pedidos_gestion_empresarial')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Gestión empresarial')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(pedido__fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None

    pedidos_por_area_articulo = pedidos.values(
        'articulo__nombre',
        'articulo__tipo'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [
        f"{item['articulo__nombre']} ({item['articulo__tipo']})"
        if item['articulo__nombre'] and item['articulo__tipo']
        else item['articulo__nombre'] or 'Sin artículo'
        for item in pedidos_por_area_articulo
    ]
    
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_gestion_empresarial.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })

@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_competitividad(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos competitividad', 'url': reverse('papeleria:pedidos_competitividad')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Competitividad')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(pedido__fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None

    pedidos_por_area_articulo = pedidos.values(
        'articulo__nombre',
        'articulo__tipo'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [
        f"{item['articulo__nombre']} ({item['articulo__tipo']})"
        if item['articulo__nombre'] and item['articulo__tipo']
        else item['articulo__nombre'] or 'Sin artículo'
        for item in pedidos_por_area_articulo
    ]
    
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_competitividad.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })

def get_pedidos_filtrados(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado__in=['Confirmado', 'Cancelado']).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio_dt, fecha_pedido__lt=fecha_fin_dt)
        except ValueError:
            pass

    return pedidos

@login_required(login_url='/acceso_denegado/')
def reporte_pedidos_pdf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Listado de pedidos CCD"
    doc.author = "CCD"
    doc.subject = "Listado de pedidos"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph("REPORTE DE PEDIDOS CONFIRMADOS Y CANCELADOS", styles["Title"])
    elements.append(titulo)

    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha"],
        ["Cámara de comercio de Duitama", "Nit: 891855025", "gestiondocumental@ccduitama.org.co", f"{fecha_actual}"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        wrap_text(usuario.username),
        wrap_text(usuario.email),
        wrap_text(getattr(usuario, 'role', 'No definido')),
        wrap_text(getattr(usuario, 'cargo', 'No definido')),
])


    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    pedidos = get_pedidos_filtrados(request).prefetch_related('articulos__articulo')

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron pedidos confirmados ni cancelados.", centered_style)
        elements.append(no_results)
    else:
        data_pedidos = [["ID", "Fecha", "Estado", "Registrado Por", "Artículos", "Área", "Devoluciones"]]

        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre} {pa.tipo}"
                for pa in pedido.articulos.all()
            ]) or 'Sin artículos'

            area_raw = pedido.articulos.first().area if pedido.articulos.exists() else 'Sin área'

            devoluciones = Devolucion.objects.filter(pedido=pedido)
            if devoluciones.exists():
                devoluciones_raw = ", ".join([
                    f"{d.cantidad_devuelta} {d.articulo.nombre}" for d in devoluciones
                ])
            else:
                devoluciones_raw = "Sin devoluciones"

            data_pedidos.append([
                wrap_text_p(str(pedido.id)),
                wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                wrap_text_p(pedido.get_estado_display()),
                wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                wrap_text_p(articulos_raw),
                wrap_text_p(area_raw),
                wrap_text_p(devoluciones_raw),
            ])

        tabla_pedidos = Table(data_pedidos, colWidths=[30, 60, 80, 120, 170, 110, 150])
        style_pedidos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_pedidos.setStyle(style_pedidos)
        elements.append(tabla_pedidos)

    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista de pedidos Gestor CCD.pdf"'
    return response

@login_required(login_url='/acceso_denegado/')
def reporte_pedidos_excel(request):
    pedidos = get_pedidos_filtrados(request).prefetch_related('articulos__articulo')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    columnas_anchos = {
        'A': 10,   
        'B': 15,   
        'C': 15,   
        'D': 25,   
        'E': 50,   
        'F': 30,   
        'G': 40 
    }
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:G1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = "Listado de Pedidos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border

    
    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Artículos', 'Áreas', 'Devoluciones']
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = "A3:G3"

    if not pedidos.exists():
        ws.merge_cells('A4:G4')
        cell = ws['A4']
        cell.value = "No hay pedidos"
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    else:
        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre} {pa.tipo}"
                for pa in pedido.articulos.all()
            ]) or 'Sin artículos'

            areas_raw = ", ".join(set([
                str(pa.area) for pa in pedido.articulos.all()
            ])) or 'Sin área'

            devoluciones_raw = ", ".join([
                f"{d.cantidad_devuelta} {d.articulo.nombre}" for d in Devolucion.objects.filter(pedido=pedido)
            ]) or 'Sin devoluciones'

            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'

            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display(),
                usuario,
                articulos_raw,
                areas_raw,
                devoluciones_raw,
            ])

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
            for i, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True if i in [4, 5, 6, 7] else False
                )

        for i in range(4, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos confirmados y cancelados.xlsx"'
    wb.save(response)
    return response

def get_pedidos_filtrados_pendientes(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado='Pendiente').order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio_dt, fecha_pedido__lt=fecha_fin_dt)
        except ValueError:
            pass  

    return pedidos

@login_required(login_url='/acceso_denegado/')
def reporte_pedidos_pendientes_pdf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Pedidos pendientes CCD"
    doc.author = "GESTOR CCD"
    doc.subject = "Pedidos pendientes CCD"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    titulo = Paragraph("Reporte de pedidos pendientes", styles["Title"])
    elements.append(titulo)

    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
       ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha"],
        ["Cámara de comercio de Duitama", "Nit: 891855025", "gestiondocumental@ccduitama.org.co", f"{fecha_actual}"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        wrap_text(usuario.username),
        wrap_text(usuario.email),
        wrap_text(getattr(usuario, 'role', 'No definido')),
        wrap_text(getattr(usuario, 'cargo', 'No definido')),
])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    pedidos = get_pedidos_filtrados_pendientes(request).prefetch_related('articulos__articulo')

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER
        )
        no_results = Paragraph("No se encontraron pedidos pendientes.", centered_style)
        elements.append(Spacer(1, 20))
        elements.append(no_results)
    else:
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Artículos", "Área"]]

        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre} {pa.tipo}"
                for pa in pedido.articulos.all()
            ]) or 'Sin artículos'
            articulos_text = wrap_text_p(articulos_raw)

            area_raw = pedido.articulos.first().area if pedido.articulos.exists() else 'Sin área'
            areas_text = wrap_text_p(area_raw)

            data_pedidos.append([
                wrap_text_p(str(pedido.id)),
                wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                wrap_text_p(pedido.get_estado_display()),
                wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                wrap_text_p(articulos_raw),
                wrap_text_p(area_raw)
            ])

        tabla_articulos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_articulos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos pendientes CCD.pdf"'
    return response

@login_required(login_url='/acceso_denegado/')
def reporte_pedidos_pendientes_excel(request):
    pedidos = get_pedidos_filtrados_pendientes(request).prefetch_related('articulos__articulo')
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    columnas_anchos = {'A': 10, 'B': 15, 'C': 15, 'D': 25, 'E': 50, 'F': 30}
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Pedidos Pendientes - {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Artículos', 'Áreas']
    ws.append(headers)
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = "A3:F3"

    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No se encontraron pedidos pendientes."
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde
    else:
        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre} {pa.tipo}"
                for pa in pedido.articulos.all()
            ])
            areas_raw = ", ".join(set(str(pa.area) for pa in pedido.articulos.all()))
            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'
            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display() if hasattr(pedido, 'get_estado_display') else pedido.estado,
                usuario,
                articulos_raw or 'Sin artículos',
                areas_raw or 'Sin área',
            ])

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for i, cell in enumerate(row, start=1):
            cell.border = borde
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True if i in [4, 5, 6] else False
            )

    for i in range(4, ws.max_row + 1):
        ws.row_dimensions[i].height = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos pendientes papeleria.xlsx"'
    wb.save(response)
    return response

#ESTADISTICAS 
@login_required(login_url='/acceso_denegado/')
def index_estadistica(request):
    breadcrumbs = [
    {'name': 'Inicio', 'url': '/index_pap'},
    {'name': 'Estadisticas', 'url': '/index_estadistica/'},  
]

    return render(request, 'estadisticas/index_estadistica.html', {'breadcrumbs': breadcrumbs})

@login_required(login_url='/acceso_denegado/')
def graficas_usuario(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadisticas', 'url': reverse('papeleria:index_estadistica')}, 
        {'name': 'Grafico de usuarios activos e inactivos', 'url': reverse('papeleria:graficas_usuarios')}, 
    ]
    activos = CustomUser.objects.filter(is_active=True).count()
    inactivos = CustomUser.objects.filter(is_active=False).count()

    nombres = ['Activos', 'Inactivos']
    cantidades = [activos, inactivos] 
    return render(request, 'estadisticas/grafica_usuarios.html', {'nombres':nombres,
                                                                  'cantidades': cantidades,
                                                                  'breadcrumbs': breadcrumbs})

@login_required(login_url='/acceso_denegado/')
def crear_devolucion(request, pedido_id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Devoluciones Papeleria', 'url': reverse('papeleria:crear_devolucion', args=[pedido_id])}
    ]
    pedido = get_object_or_404(Pedido, id=pedido_id)

    # Validación: no permitir devoluciones si el pedido está cancelado
    if pedido.estado == "Cancelado":
        messages.error(request, "No puedes registrar una devolución porque el pedido ha sido cancelado.")
        return redirect('papeleria:mis_pedidos')

    if request.method == 'POST':
        form = DevolucionForm(request.POST, pedido_id=pedido_id)
        if form.is_valid():
            devolucion = form.save(commit=False)
            articulo = devolucion.articulo
            cantidad_devuelta = devolucion.cantidad_devuelta

            try:
                pedido_articulo = PedidoArticulo.objects.get(pedido=pedido, articulo=articulo)
            except PedidoArticulo.DoesNotExist:
                messages.error(request, "El artículo no pertenece a este pedido.")
                return render(request, 'pedidos/devolver_articulo.html', {'form': form, 'pedido': pedido})

            total_devuelto = Devolucion.objects.filter(
                pedido=pedido, articulo=articulo
            ).aggregate(total=Sum('cantidad_devuelta'))['total'] or 0

            if total_devuelto >= pedido_articulo.cantidad:
                messages.error(request, "Ya has devuelto este artículo en su totalidad.")
            elif total_devuelto + cantidad_devuelta > pedido_articulo.cantidad:
                restante = pedido_articulo.cantidad - total_devuelto
                messages.error(request, f"Solo puedes devolver hasta {restante} unidades.")
            else:
                devolucion.save()
                messages.success(request, "La devolución fue registrada exitosamente.")
                return redirect('papeleria:mis_pedidos')
        else:
            messages.error(request, "Verifica que los campos sean válidos.")
    else:
        form = DevolucionForm(pedido_id=pedido_id)

    return render(request, 'pedidos/devolver_articulo.html', {
        'form': form,
        'pedido': pedido,
        'breadcrumbs': breadcrumbs
    })